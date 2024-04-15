# -*- coding: utf-8 -*-
import csv
import shutil
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font
import os
from pathlib import Path

import logging

logging.basicConfig(level=logging.INFO,
                    format='%(levelname)s - %(asctime)s  - %(message)s',
                    datefmt='%H:%M:%S')

# Стили
font = Font(name='Times New Roman', size=9)
font_bold = Font(name='Times New Roman', size=9, bold=True)

alignment = Alignment(horizontal='center',
                      vertical='center',
                      wrap_text=True
                      )
alignment_2 = Alignment(horizontal='left',
                        vertical='center',
                        wrap_text=True
                        )
alignment_3 = Alignment(horizontal='left',
                        vertical='top',
                        wrap_text=True
                        )
border = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))


def check_months(m1s, m1e, m2s, m2e, m3s, m3e):
    """Обработка количества месяцев."""
    # Создаем список месяцев
    months_list = [m1s, m1e, m2s, m2e, m3s, m3e]

    # Фильтруем пустые элементы
    valid_months = [elem for elem in months_list if elem != '-']

    # Вычисляем количество пустых элементов
    delta = max(0, 6 - len(valid_months))

    # Дополняем список пустыми значениями
    result = valid_months + ['-' for _ in range(delta)]

    # Добавляем количество заполненных элементов в результат
    result.append(int(len(valid_months) / 2))

    return result


def fill_month_data(sheet, idx_row, i, month_start, month_end):
    """Заполнение строк датами, '-' и 1."""
    if month_start != '-':
        sheet[f'D{idx_row + i}'] = month_start
        sheet[f'E{idx_row + i}'] = month_end
        sheet[f'F{idx_row + i}'] = f'=E{idx_row + i}-D{idx_row + i}+1'

        for column in range(7, 13):
            sheet.cell(row=idx_row + i, column=column, value='−')
        sheet[f'M{idx_row + i}'] = 1


def fill_signature(sheet, idx_row, signature):
    """Создание подписи в документе."""
    sheet[f'A{idx_row + 1}'] = 'Подписи сторон:'
    sheet[f'A{idx_row + 1}'].font = font
    sheet[f'A{idx_row + 1}'].alignment = alignment
    sheet.merge_cells(f'A{idx_row + 1}:M{idx_row + 1}')

    sheet[f'B{idx_row + 3}'] = 'ИСПОЛНИТЕЛЬ:'
    sheet[f'B{idx_row + 3}'].font = font_bold

    sheet[f'J{idx_row + 3}'] = 'МО:'
    sheet[f'J{idx_row + 3}'].font = font_bold

    sheet[f'B{idx_row + 5}'] = 'ООО «Эврика»\nМенеджер'
    sheet[f'B{idx_row + 5}'].font = font_bold
    sheet[f'B{idx_row + 5}'].alignment = alignment_2
    sheet.merge_cells(f'B{idx_row + 5}:D{idx_row + 5}')
    sheet.row_dimensions[idx_row + 5].height = 30

    sheet[f'J{idx_row + 5}'] = signature
    sheet[f'J{idx_row + 5}'].font = font_bold
    sheet[f'J{idx_row + 5}'].alignment = alignment_3
    sheet.merge_cells(f'J{idx_row + 5}:L{idx_row + 5}')

    sheet[f'B{idx_row + 8}'] = '________________/А.А. Шубенков/'
    sheet[f'B{idx_row + 8}'].font = font

    sheet[f'J{idx_row + 8}'] = '_________________ /___________________'

    for i in ('B', 'J'):
        sheet[f'{i}{idx_row + 9}'] = 'М.П.'
        sheet[f'{i}{idx_row + 9}'].font = font


def table_style(sheet):
    """Оформление таблицы."""
    for row in sheet.iter_rows(values_only=False):
        for cell in row:
            cell.border = border
            cell.font = font
            cell.alignment = alignment

            if isinstance(cell.value, str):
                if 'государственное' in cell.value.lower():
                    cell.font = font_bold
                    cell.alignment = alignment_2
                elif 'место' in cell.value.lower():
                    cell.alignment = alignment_2

    # Сквозные строки для таблицы
    sheet.print_title_rows = '1:2'


def file_processing(sheet, data):
    """Создание Актов из CSV файлов."""
    number = 0
    cnt_row = None
    idx_row = 3
    current_address = None
    current_mo = None
    for row in data:
        point = row['ТТ']
        type_point = row['Тип']
        month_1_start = row['н1']
        month_1_end = row['к1']
        month_2_start = row['н2']
        month_2_end = row['к2']
        month_3_start = row['н3']
        month_3_end = row['к3']
        signature = row['Подпись']
        mo = row['Наименование МО']
        pref = 'Место оказания услуги: '
        address = row['Адрес']

        months = [month_1_start, month_1_end, month_2_start,
                  month_2_end, month_3_start, month_3_end]

        month_1_start, month_1_end, \
        month_2_start, month_2_end, \
        month_3_start, month_3_end, cnt_row = check_months(*months)

        if current_address != address or current_mo != mo:
            sheet[f'A{idx_row}'] = mo
            sheet[f'A{idx_row + 1}'] = pref + address

            sheet.merge_cells(f'A{idx_row}:M{idx_row}')
            sheet.row_dimensions[idx_row].height = 30

            sheet.merge_cells(f'A{idx_row + 1}:M{idx_row + 1}')
            sheet.row_dimensions[idx_row + 1].height = 15

            current_address = row['Адрес']
            current_mo = row['Наименование МО']

            idx_row += 2
            number = 0

        sheet[f'A{idx_row}'] = number + 1
        sheet[f'B{idx_row}'] = point
        sheet[f'C{idx_row}'] = type_point

        fill_month_data(sheet, idx_row, 0, month_1_start, month_1_end)
        fill_month_data(sheet, idx_row, 1, month_2_start, month_2_end)
        fill_month_data(sheet, idx_row, 2, month_3_start, month_3_end)

        # Объединение ячеек в таблице
        for i in ('A', 'B', 'C'):
            sheet.merge_cells(f'{i}{idx_row}:{i}{idx_row + cnt_row - 1}')

        idx_row += cnt_row
        number += 1

    # Делаем оформление для таблицы
    table_style(sheet)

    # Создание подписи в конце документа
    fill_signature(sheet, idx_row, signature)


def create_csv(template):
    """
    Разбивка данных на CSV файлы.

        Колонки:
        - Технологическая точка;
        - Тип;
        - Код МО;
        - Наименование МО;
        - Начало 1-го месяца;
        - Конец 1-го месяца;
        - Начало 2-го месяца;
        - Конец 2-го месяца;
        - Начало 3-го месяца;
        - Конец 3-го месяца;
        - Подпись для МО;
        - Общее МО.
    """
    with open(template, newline='') as csvfile:
        reader = csv.DictReader(csvfile, delimiter=';')

        current_sign = None
        current_file = None

        for row in reader:
            sign = row['Подпись']
            name_file = row['Общее МО']
            if sign != current_sign:
                current_sign = sign
                if current_file:
                    current_file.close()

                filename = 'csv/' + name_file + '.csv'

                current_file = open(
                    filename, 'a',
                    newline='',
                    encoding='utf-8'
                )
                fieldnames = reader.fieldnames
                writer = csv.DictWriter(current_file, fieldnames=fieldnames)
                if current_file.tell() == 0:
                    writer.writeheader()
            writer.writerow(row)

        if current_file:
            current_file.close()


def get_new_file_name(file_name, folder):
    """Переименовывает новый файл если уже есть такой фал в папке."""
    base, ext = os.path.splitext(file_name)
    index = 1
    new_file_name = file_name

    while os.path.exists(os.path.join(folder, new_file_name)):
        new_file_name = f"{base} ({index}){ext}"
        index += 1
    return new_file_name


def main():
    num = input('Укажите номер очереди: ')
    template = 'Шаблон_для_акта.xlsx'
    folder_name = f'Акты_{num}-я_очередь'
    template_date = f'Данные_для_актов/excel_data_{num}.csv'

    if not os.path.exists(template_date):
        logging.warning(f'Нет файла "{template_date}" для "{num}" очереди!')
        exit()

    if not os.path.exists(folder_name):
        os.mkdir(folder_name)
        logging.info(f'Создаю папку "{folder_name}"')

    # Удаляем папку "csv" если она есть
    if os.path.exists('csv'):
        shutil.rmtree('csv')

    # Создаем папку "csv" если ее нет
    if not os.path.exists('csv'):
        os.mkdir('csv')

    # Разделяем данные на отдельные файлы csv
    create_csv(template_date)

    # Счетчик файлов
    quantity = 0

    csv_folder = 'csv'
    csv_files = os.listdir(csv_folder)

    for file in csv_files:
        with open(os.path.join(csv_folder, file),
                  newline='', encoding='utf-8') as csvfile:
            data_data = csv.DictReader(csvfile, delimiter=",")

            name_file = Path(file).stem
            logging.info(f'Создаю файл "{name_file}"')

            # Открываем шаблон
            folder = 'Шаблоны'
            wb = openpyxl.load_workbook(f'{folder}/{template}')

            # Подгружаем лист в переменную
            sheet = wb['Лист1']

            # Создаем акт
            file_processing(sheet, data_data)

            # Присвоение имени для документа
            file = get_new_file_name(name_file + '.xlsx', folder_name)

            with Path(folder_name, file) as output_file:
                wb.save(output_file)
                quantity += 1
                logging.info(f'{name_file} - [ok]')

    # Удаляем папку "csv"
    shutil.rmtree('csv')

    logging.info(f'Создано файлов: {quantity}')
    logging.info('Операция завершена успешно')


if __name__ == "__main__":
    main()
