import csv
from constants import (SIGNATURE_EXECUTIVE, SIGNATURES_PARTIES, EXECUTIVE,
                       EURECA, SIGNATURE_MEDICAL_ORGANIZATION,
                       MEDICAL_ORGANIZATION, STAMP_PLACE)
import shutil
from pprint import pprint
import subprocess
from styles import (font, font_bold, border,
                    alignment_1, alignment_2, alignment_3)
import openpyxl
import os
from pathlib import Path

import logging

logging.basicConfig(level=logging.INFO,
                    format='%(levelname)s - %(asctime)s  - %(message)s',
                    datefmt='%H:%M:%S')


def check_months(*args):
    """Обработка количества месяцев."""

    result = ['-'] * 6
    idx = 0
    for month in args:
        if month != '-':
            result[idx] = month
            idx += 1

    idx_month = str(idx // 2)    # Добавление количества месяцев
    result.append(idx_month)
    return result


def fill_month_data(sheet, idx_row, i, month_start, month_end):
    """Заполнение строк датами, '-' и 1."""
    if month_start != '-':
        sheet[f'D{idx_row + i}'] = month_start
        sheet[f'E{idx_row + i}'] = month_end
        sheet[f'F{idx_row + i}'] = f'=E{idx_row + i}-D{idx_row + i}+1'

        for column in range(7, 13):
            sheet.cell(row=idx_row + i, column=column, value='−')
        sheet[f'M{idx_row + i}'] = 1


def fill_signature(sheet, idx_row, signature):
    """Создание подписи в документе."""
    sheet[f'A{idx_row + 1}'] = SIGNATURES_PARTIES
    sheet[f'A{idx_row + 1}'].font = font
    sheet[f'A{idx_row + 1}'].alignment = alignment_1
    sheet.merge_cells(f'A{idx_row + 1}:M{idx_row + 1}')

    sheet[f'B{idx_row + 3}'] = EXECUTIVE
    sheet[f'B{idx_row + 3}'].font = font_bold

    sheet[f'J{idx_row + 3}'] = MEDICAL_ORGANIZATION
    sheet[f'J{idx_row + 3}'].font = font_bold

    sheet[f'B{idx_row + 5}'] = EURECA
    sheet[f'B{idx_row + 5}'].font = font_bold
    sheet[f'B{idx_row + 5}'].alignment = alignment_2
    sheet.merge_cells(f'B{idx_row + 5}:D{idx_row + 5}')
    sheet.row_dimensions[idx_row + 5].height = 30

    sheet[f'J{idx_row + 5}'] = signature
    sheet[f'J{idx_row + 5}'].font = font_bold
    sheet[f'J{idx_row + 5}'].alignment = alignment_3
    sheet.merge_cells(f'J{idx_row + 5}:L{idx_row + 5}')

    sheet[f'B{idx_row + 8}'] = SIGNATURE_EXECUTIVE
    sheet[f'B{idx_row + 8}'].font = font

    sheet[f'J{idx_row + 8}'] = SIGNATURE_MEDICAL_ORGANIZATION

    for i in ('B', 'J'):
        sheet[f'{i}{idx_row + 9}'] = STAMP_PLACE
        sheet[f'{i}{idx_row + 9}'].font = font


def table_style(sheet):
    """Оформление таблицы."""
    for row in sheet.iter_rows(values_only=False):
        for cell in row:
            cell.border = border
            cell.font = font
            cell.alignment = alignment_1

            if isinstance(cell.value, str):
                if 'государственное' in cell.value.lower():
                    cell.font = font_bold
                    cell.alignment = alignment_2
                elif 'место' in cell.value.lower():
                    cell.alignment = alignment_2

    sheet.print_title_rows = '1:2'  # Сквозные строки для таблицы


def file_processing(template, folder, csvfile):
    """Создание Актов из CSV файлов."""
    number = 0
    idx_row = 3
    current_address = None
    current_mo = None
    current_file = None
    wb = None
    current_signature = None
    quantity = 0

    with open(csvfile, newline='') as csvfile:
        reader = csv.DictReader(csvfile, delimiter=';')
        for row in reader:
            point = row['ТТ']
            type_point = row['Тип']
            m_1_start = row['н1']
            m_1_end = row['к1']
            m_2_start = row['н2']
            m_2_end = row['к2']
            m_3_start = row['н3']
            m_3_end = row['к3']
            signature = row['Подпись']
            mo = row['Наименование МО']
            pref = 'Место оказания услуги: '
            address = row['Адрес']
            file = row['Общее МО']

            months = [m_1_start, m_1_end, m_2_start, m_2_end, m_3_start, m_3_end]

            (m_1_start, m_1_end,
             m_2_start, m_2_end,
             m_3_start, m_3_end,
             cnt_row) = check_months(*months)

            cnt_row = int(cnt_row)

            if current_file != file:
                if wb is not None:
                    table_style(sheet)
                    fill_signature(sheet, idx_row, current_signature)
                    new_f_name = get_new_file_name(
                        f'{current_file}.xlsx',
                        folder
                    )
                    with Path(folder, new_f_name) as output_file:
                        wb.save(output_file)
                    quantity += 1
                    logging.info(f'[+] {current_file}')

                wb = openpyxl.load_workbook(template)
                sheet = wb['Лист1']
                current_file = file
                idx_row = 3

            if current_address != address or current_mo != mo:
                sheet[f'A{idx_row}'] = mo
                sheet[f'A{idx_row + 1}'] = pref + address

                sheet.merge_cells(f'A{idx_row}:M{idx_row}')
                sheet.row_dimensions[idx_row].height = 30

                sheet.merge_cells(f'A{idx_row + 1}:M{idx_row + 1}')
                sheet.row_dimensions[idx_row + 1].height = 15

                current_address = row['Адрес']
                current_mo = row['Наименование МО']

                idx_row += 2
                number = 0

            sheet[f'A{idx_row}'] = number + 1
            sheet[f'B{idx_row}'] = point
            sheet[f'C{idx_row}'] = type_point

            fill_month_data(sheet, idx_row, 0, m_1_start, m_1_end)
            fill_month_data(sheet, idx_row, 1, m_2_start, m_2_end)
            fill_month_data(sheet, idx_row, 2, m_3_start, m_3_end)

            for i in ('A', 'B', 'C'):  # Объединение ячеек в таблице
                sheet.merge_cells(f'{i}{idx_row}:{i}{idx_row + cnt_row - 1}')

            idx_row += cnt_row
            number += 1
            current_signature = signature

        table_style(sheet)
        fill_signature(sheet, idx_row, current_signature)

        if current_file is not None:
            new_f_name = get_new_file_name(f'{current_file}.xlsx', folder)
            with Path(folder, new_f_name) as output_file:
                wb.save(output_file)
            quantity += 1
            logging.info(f'[+] {current_file}')
    logging.info(f'Создано файлов: {quantity}')


def get_new_file_name(file_name, folder):
    """Переименовывает новый файл если уже есть такой файл в папке."""
    base, ext = os.path.splitext(file_name)
    index = 1
    new_file_name = file_name

    while os.path.exists(os.path.join(folder, new_file_name)):
        new_file_name = f'{base} ({index}){ext}'
        index += 1
    return new_file_name


def check_queue_num():
    """Проверка корректности очереди."""
    while True:
        queue_num = input('Укажите номер очереди (1, 4, 5 или 13) '
                          'или 0 для выхода: ')
        if queue_num in ('1', '4', '5', '13'):
            return queue_num
        if queue_num == '0':
            exit('Завершение работы')
        else:
            print('Необходимо указать правильно очередь!')


def main():
    """Главная функция для создания технических актов.

    Имена файлов с данными должны иметь вид excel_data_{номер очереди}.csv
    """

    print('Выбран режим создания Технических актов')
    queue_num = check_queue_num()
    template_path = fr'Шаблоны\Шаблон_для_акта.xlsx'
    folder_name = f'Акты_{queue_num}-я_очередь'
    csv_data = fr'Данные_для_актов\excel_data_{queue_num}.csv'

    if not os.path.exists(csv_data):
        exit(f'Нет файла "{csv_data}" для "{queue_num}" очереди!')

    if not os.path.exists(folder_name):
        os.mkdir(folder_name)
        logging.info(f'Создаю папку "{folder_name}"')

    file_processing(template_path, folder_name, csv_data)

    logging.info('Операция завершена успешно')

    project_folder = os.getcwd()
    subprocess.Popen(fr'explorer {project_folder}\{folder_name}')


if __name__ == "__main__":
    main()
