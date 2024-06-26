from constants import M1, M2, M3
import subprocess

import openpyxl
import os
from pathlib import Path
import re
import csv
import logging


logging.basicConfig(level=logging.INFO,
                    format='%(levelname)s - %(asctime)s  - %(message)s',
                    datefmt='%H:%M:%S')


def check_month(m1s, m1e, m2s='-', m2e='-', m3s='-', m3e='-'):
    """
    Распределение дат по своим месяцам.

    Необходимо менять ключи под нужный этап в файле constants.
    """
    months = (m1s, m1e, m2s, m2e, m3s, m3e)
    month_mapping = {M1: (0, 1), M2: (2, 3), M3: (4, 5)}
    result = ['-' for _ in range(6)]

    for i in range(0, len(months), 2):
        month = months[i]
        if month and month[3:5] in month_mapping:
            start_idx, end_idx = month_mapping[month[3:5]]
            result[start_idx] = month
            result[end_idx] = months[i + 1]

    return result


def write_to_csv(data, file_name):
    """Создание CSV файла и внесение данных."""
    file_exists = os.path.isfile(file_name)
    with open(file_name, mode='a', newline='') as file:
        writer = csv.writer(file, delimiter=';')

        if not file_exists:
            writer.writerow(
                ['ТТ', 'Тип', 'Наименование МО', 'Адрес',
                 'н1', 'к1', 'н2', 'к2', 'н3', 'к3', 'Подпись', 'Общее МО']
            )

        writer.writerow(data)


def file_processing(sheet, file, num):
    """Просмотр файла."""
    title = None
    address = None

    csv_file = f'Данные (СП{num}).csv'
    csv_folder = 'Анализ_актов'
    csv_path = f'{csv_folder}/{csv_file}'

    logging.info(f'Обработка "{file}"')
    for ir in range(1, sheet.max_row + 1):
        for ic in range(1, 3):
            obj = str(sheet.cell(ir, ic).value)
            if re.search(r'\*\d{3}\-\d{4}\*', obj):
                point = obj
                type_point = sheet[f'C{ir}'].value
                month_1_start = sheet[f'D{ir}'].value
                month_1_end = sheet[f'E{ir}'].value
                month_2_start, month_2_end = "-", "-"
                month_3_start, month_3_end = "-", "-"

                if 'учреждение' in str(sheet[f'A{ir - 2}'].value):
                    title = sheet[f'A{ir - 2}'].value

                if 'услуги:' in str(sheet[f'A{ir - 1}'].value):
                    address = sheet[f'A{ir - 1}'].value[23:]

                if (not sheet.cell(ir + 1, 2).value
                        and sheet.cell(ir + 1, column=4).value):
                    month_2_start = sheet[f'D{ir + 1}'].value
                    month_2_end = sheet[f'E{ir + 1}'].value

                if (not sheet.cell(ir + 2, 1).value
                        and sheet.cell(ir + 2, column=4).value
                        and month_2_start != "-"):
                    month_3_start = sheet[f'D{ir + 2}'].value
                    month_3_end = sheet[f'E{ir + 2}'].value

                signature = sheet[f'J{sheet.max_row - 4}'].value
                months = [month_1_start, month_1_end,
                          month_2_start, month_2_end,
                          month_3_start, month_3_end]
                months = check_month(*months)
                data_to_write = (point, type_point, title, address,
                                 *months, signature, file)

                write_to_csv(data_to_write, csv_path)


def main():
    """Главная функция."""
    print('Выбран режим сбора данных из актов')
    num = input('Укажите номер очереди (1, 4, 5 или 13): ')

    if num not in ('1', '4', '5', '13'):
        exit('Необходимо указать правильно очередь!')

    folder_name = fr'Анализ_актов\input\{num}-я'
    files = os.listdir(folder_name)

    if not files:
        exit(f'Нет файлов в папке "{folder_name}" для обработки!')

    for file in files:
        file_name = Path(file).stem
        if Path(file).suffix == '.xlsx':
            file_path = Path(folder_name, file)
            wb = openpyxl.load_workbook(file_path)
            wb_sheet = wb['Лист1']
            file_processing(wb_sheet, file_name, num)
    logging.info(f'Анализ завершен. Создан файл "Данные (СП{num}).csv"')
    project_folder = os.getcwd()
    subprocess.Popen(fr'explorer {project_folder}\Анализ_актов')


if __name__ == "__main__":
    main()
