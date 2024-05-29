import csv
from datetime import datetime, timedelta
import os
from pathlib import Path
import re
import subprocess
import logging

import openpyxl


logging.basicConfig(level=logging.INFO,
                    format='%(levelname)s - %(asctime)s  - %(message)s',
                    datefmt='%H:%M:%S')

incidents = []
points = []


def create_list_incident(data):
    """Функция для создания листа с заявками из CSV файла."""
    with open(data, newline='') as csvfile:
        rows = csv.DictReader(csvfile, delimiter=';')
        for row in rows:
            incidents.append(row)
            points.append(row['ТТ'])


def create_acts_with_incident(acts, folder_input, folder_output):
    """Функция для создания актов с заявками."""
    quantity = 0

    for act in acts:
        logging.info(f'Открываю {Path(act).stem}')
        if Path(act).suffix == '.xlsx':
            file_path = Path(folder_input, act)
            wb = openpyxl.load_workbook(file_path)
            sheet = wb['Лист1']
            find_point(sheet)
            sheet.print_title_rows = '1:2'  # Скозные строки

            # Сохранение файла в папку
            with Path(folder_output, act) as output_file:
                wb.save(output_file)
                logging.info(f'{Path(act).stem} - [+]')
                quantity += 1

    logging.info(f'Операция завершена успешно. '
                 f'Всего обработано файлов - {quantity}')


def str_to_date(string, flag=False):
    """Функция преобразования даты."""
    if len(string) == 10:
        date_obj = datetime.strptime(string, '%d.%m.%Y')
        if flag:
            date_obj += timedelta(hours=23, minutes=59)
        return date_obj
    date_obj = datetime.strptime(string, '%d.%m.%Y %H:%M')
    return date_obj


def fill_incident(sheet, idx, start, end, point):
    """Функция добавления заявок."""
    if incidents:
        for i, row in enumerate(incidents):
            if (
                (row['ТТ'] == point and
                 start <= str_to_date(row['Время назначения']) <= end)
            ):
                sheet[f'G{idx}'] = row['Номер заявки']
                sheet[f'H{idx}'] = row['Время в отложено']
                sheet[f'I{idx}'] = row['Время обработки']
                sheet[f'J{idx}'] = row['Время назначения']
                sheet[f'K{idx}'] = row['Время закрытия']
                sheet[f'L{idx}'] = row['Время ограничения']
                sheet[f'M{idx}'] = int(row['Коэффициент'])
                incidents.pop(i)


def check_month(sheet, point, idx):
    """Функция проверки наличия 2-го и 3-го месяца."""
    if not sheet[f'B{idx}'].value and sheet[f'D{idx}'].value:
        month_start = str_to_date(sheet[f'D{idx}'].value)
        month_end = str_to_date(sheet[f'E{idx}'].value, True)
        fill_incident(sheet, idx, month_start, month_end, point)


def find_point(sheet):
    """Функция поиска технологических точек."""
    pattern_point = r'\*\d{3}\-\d{4}\*'

    for ir in range(1, sheet.max_row + 1):
        idx_2 = ir + 1
        idx_3 = ir + 2

        for ic in range(1, 3):
            obj = str(sheet.cell(ir, ic).value)
            point = obj

            if re.search(pattern_point, obj) and point in points:

                # 1-й месяц
                month_1_start = str_to_date(sheet[f'D{ir}'].value)
                month_1_end = str_to_date(sheet[f'E{ir}'].value, True)
                fill_incident(sheet, ir, month_1_start, month_1_end, point)

                # 2-й месяц
                check_month(sheet, point, idx_2)

                # 3-й месяц
                check_month(sheet, point, idx_3)


def main():
    """Главная функция."""
    print('Выбран режим внесения инцидентов в Технические акты')
    num = input('Укажите номер очереди (1, 4, 5 или 13): ')

    if num not in ('1', '4', '5', '13'):
        exit('Необходимо указать правильно очередь!')

    folder_input = fr'Акты_с_заявками\input\{num}-я'
    folder_output = fr'Акты_с_заявками\output\{num}-я'
    data = fr'Данные_для_актов\incident_{num}.csv'
    acts = os.listdir(folder_input)

    if not os.path.exists(data):
        exit(f'Нет файла "{data}" для "{num}" очереди!')

    create_list_incident(data)
    create_acts_with_incident(acts, folder_input, folder_output)
    logging.info('Открываю каталог с результатом')
    project_folder = os.getcwd()
    subprocess.Popen(fr'explorer {project_folder}\{folder_output}')


if __name__ == "__main__":
    main()
